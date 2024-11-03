import os, re, pathlib, datetime, time, copy
import sys
import openpyxl

def read_dir(name_dir: str) -> list:
    """
    Считывает файлы из папки 
    """
    list_files = []
    for file in os.listdir(name_dir):
        filenames = ''.join(c for c in file if c.isprintable())
        try:
            os.rename(file, filenames)
        except FileNotFoundError:
            pass
        list_files.append(file)
    return list_files

def data_file(filenames):
    """
    Функция даты файлов
    """
    ti_m = os.path.getmtime(filenames)
    m_ti = time.ctime(ti_m)
    t_obj = time.strptime(m_ti)
    T_stamp = time.strftime("%Y.%m.%d", t_obj)
    return T_stamp

def pattern_(file: str):
    """Файл PDFный???"""
    pdf_ = re.fullmatch(r'.*.pdf', file)
    dwg_ = re.fullmatch(r'.*.dwg', file)
    exe_ = re.fullmatch(r'.*.exe', file)
    png_ = re.fullmatch(r'.*.png', file)
    jpg_ = re.fullmatch(r'.*.jpg', file)
    tiff_ =re.fullmatch(r'.*.tiff', file)
    if pdf_ or exe_ :
        return True
    
def pattern_all(file: str):
    """Файл PDFный???"""
    pdf_ = re.fullmatch(r'.*.pdf', file)
    dwg_ = re.fullmatch(r'.*.dwg', file)
    exe_ = re.fullmatch(r'.*.exe', file)
    png_ = re.fullmatch(r'.*.png', file)
    jpg_ = re.fullmatch(r'.*.jpg', file)
    tiff_ =re.fullmatch(r'.*.tiff', file)
    if dwg_ or pdf_ or exe_ or png_ or jpg_ or tiff_:
        return True

template_file = 'template.xlsx'
files_sheet_name = 'files'
filenames = 'filenames'

directory = os.path.abspath(os.getcwd())
directory_dwg = os.path.join(directory, "DWG")
directory_pdf = os.path.join(directory, "PDF")
#list_dir = [directory_pdf, directory_dwg, directory]

from_PDF = read_dir(directory_pdf) #Файлы из папки PDF
from_DWG = read_dir(directory_dwg) #Файлы из папки DWG

# 1. Получаем список всех файлов в текущей папке
files = [f[:-4] for f in from_DWG if f.endswith('.pdf') or f.endswith('.dwg')]
files = list(set(files))
files.sort()

# 2. Открываем Excel файл и записываем имена PDF файлов в лист 'files'
wb = openpyxl.load_workbook(template_file)
ws_files = wb[files_sheet_name]

# Очищаем лист перед записью новых данных
ws_files.delete_rows(1, ws_files.max_row)

for index, file in enumerate(files, start=1):
    ws_files.cell(row=index, column=1, value=file)

wb.save(template_file)

# 3. Читаем имена файлов с листа 'filenames'
ws_sort = wb[filenames]
right_name = [cell.value for cell in ws_sort['A'] if cell.value is not None]
export1 = [cell.value for cell in ws_sort['B'] if cell.value is not None]
export2 = [cell.value for cell in ws_sort['C'] if cell.value is not None]
export3 = [cell.value for cell in ws_sort['D'] if cell.value is not None]
export4 = [cell.value for cell in ws_sort['E'] if cell.value is not None]
export5 = [cell.value for cell in ws_sort['F'] if cell.value is not None]

#Проверяем валидность записанных значений
if len(right_name) != len(export1):
    print('Проверьте столбец B')
    time.sleep(10)
    sys.exit()
elif len(right_name) != len(export2) and len(export2) != 0:
    print('Проверьте столбец C')
    time.sleep(10)
    sys.exit()
elif len(right_name) != len(export3) and len(export3) != 0:
    print('Проверьте столбец D')
    time.sleep(10)
    sys.exit()
elif len(right_name) != len(export4) and len(export4) != 0:
    print('Проверьте столбец E')
    time.sleep(10)
    sys.exit()
elif len(right_name) != len(export5) and len(export5) != 0:
    print('Проверьте столбец F')
    time.sleep(10)
    sys.exit()

new_dict_name = {}
for i, name in enumerate(right_name):
    # if len(right_name)==len(export1)==len(export2)==len(export3):
    if export1 and export2 and export3 and export4 and export5:
        try:
            new_dict_name[name] = export1[i], export2[i], export3[i], export4[i], export5[i]
        except IndexError:
            print('Заполните все ячейки в столбце')
            time.sleep(10)
        else:
            new_dict_name[name] = export1[i], export2[i], export3[i], export4[i], export5[i]
    elif export1 and export2 and export3 and export4:
        try:
            new_dict_name[name] = export1[i], export2[i], export3[i], export4[i]
        except IndexError:
            print('Заполните все ячейки в столбце')
            time.sleep(10)
        else:
            new_dict_name[name] = export1[i], export2[i], export3[i], export4[i]
    elif export1 and export2 and export3:
        try:
            new_dict_name[name] = export1[i], export2[i], export3[i]
        except IndexError:
            print('Заполните все ячейки в столбце')
            time.sleep(10)
        else:
            new_dict_name[name] = export1[i], export2[i], export3[i]
    elif export1 and export2:
        try:
            new_dict_name[name] = export1[i], export2[i]
        except IndexError:
            print('Заполните все ячейки в столбце')
            time.sleep(10)
        else:
            new_dict_name[name] = export1[i], export2[i]
    elif export1:
        try:
            new_dict_name[name] = export1[i]
        except IndexError:
            print('Заполните все ячейки в столбце')
            time.sleep(10)
        else:
            new_dict_name[name] = export1[i]


#auto-py-to-exe
#T:\Проектный институт\Группа архитекторов\Группа АР БКП\LETV-01\05_ПД\Раздел 3.  Объемно-планировочные и архитектурные решения\
def replace_file(full_filename_now,full_filename_new):
    """
    Функция переноса файлов
    """
    file = os.path.basename(full_filename_now)
    time_ = datetime.datetime.now().replace(microsecond=0)
    dir_now = os.path.basename(os.path.dirname(full_filename_now))
    dir_new = os.path.basename(os.path.dirname(full_filename_new))
    print(f"Перенесли     {dir_now}:{file[:5]}...{file[-30:]} --> в папку {dir_new}. Время: {time_}")
    os.replace(full_filename_now, full_filename_new)
    
def dict_right_filename(find_name):
    """Подбирает к неправильному имени правильное"""
    strip_name, file_extension = os.path.splitext(find_name)
    if pattern_all(find_name):
        for keys, values in new_dict_name.items():
            if values[1] == strip_name:
                return keys + file_extension
            elif values[0] == strip_name:
                return keys + file_extension
            else:
                pass


           
def main():
    directory = os.path.abspath(os.getcwd())
    directory_dwg = os.path.join(directory, "DWG")
    directory_pdf = os.path.join(directory, "PDF")
    #list_dir = [directory_pdf, directory_dwg, directory]

    from_PDF = read_dir(directory_pdf) #Файлы из папки PDF
    from_DWG = read_dir(directory_dwg) #Файлы из папки DWG
    # print("проверка ПДФ")
    for file in from_DWG:
        on_rename = dict_right_filename(file)
        full_file = os.path.join("DWG", file)
        if on_rename != None:
            full_on_rename = os.path.join("DWG", on_rename)
            if on_rename in from_DWG:
                print(f"Удален        {on_rename}")
                os.remove(full_on_rename)

        
            
    for file in from_DWG:
        on_rename = dict_right_filename(file)
        full_file = os.path.join("DWG", file)
        if on_rename != None:
            full_on_rename = os.path.join("DWG", on_rename)
            print(f"Переименовали {file} --> {on_rename[:25]}...{on_rename[-30:]}")
            os.rename(full_file, full_on_rename)
            # print(f"Переименовка True")


  
    for file in from_DWG:
        full_name = os.path.join("DWG", file)
        full_name_new = os.path.join("PDF", file)
        #Ищем НЕ PDF файлы, после чего запускаем удаление/перемещение
     
        if pattern_(file):
            replace_file(full_name, full_name_new)


while(True):
    # main()
    try:
        main()
    except Exception:
        time.sleep(5)
    time.sleep(5)

